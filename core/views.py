from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q
from django.http import JsonResponse
from .models import Item, Transaccion, Categoria, Subcategoria
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.contrib import messages # Para enviar alertas al usuario
from datetime import datetime
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import PermissionDenied
from django.shortcuts import render
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from .models import Item, Categoria, Subcategoria


def error_403(request, exception=None): # El '=None' es vital
    try:
        return render(request, 'core/403.html', status=403)
    except Exception as e:
        # Esto te dirá en la consola del servidor qué falló exactamente
        print(f"Error renderizando 403: {e}")
        from django.http import HttpResponse
        return HttpResponse("Error en el template 403", status=403)

def error_404(request, exception):
    return render(request, 'core/404.html', status=404)

# Create your views here.
@login_required
def lista_inventario(request):
    busqueda = request.GET.get('buscar') # Captura lo que el usuario escribe
    items = Item.objects.all()

    if busqueda:
        items = items.filter(
            Q(nombre__icontains=busqueda) | 
            Q(sku__icontains=busqueda)
        )

    return render(request, 'core/lista_inventario.html', {'items': items, 'busqueda': busqueda})

# Función para el botón de salida rápida (-1)
@login_required
def salida_rapida(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if item.stock > 0:
        item.stock -= 1
        item.save() # Esto dispara automáticamente la señal que creamos antes
    return redirect('lista_inventario')

# Vista para ver todos los movimientos
@login_required
def historial_movimientos(request):
    movimientos = Transaccion.objects.all().order_by('-fecha')
    return render(request, 'core/historial.html', {'movimientos': movimientos})

from django.shortcuts import render, get_object_or_404, redirect
from .models import Item, Transaccion
from django import forms

class SubcategoriaChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        return f"{obj.nombre}" # Forzamos que devuelva solo el nombre

class CategoriaChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        return f"{obj.nombre}"

# Creamos un formulario rápido para editar el ítem
class ItemForm(forms.ModelForm):
    class Meta:
        model = Item
        fields = ['nombre', 'sku', 'stock', 'stock_minimo', 'descripcion', 'imagen'] # Agregados aquí
        widgets = {
            'nombre': forms.TextInput(attrs={'class': 'form-control'}),
            'sku': forms.TextInput(attrs={'class': 'form-control'}),
            'stock': forms.NumberInput(attrs={'class': 'form-control'}),
            'stock_minimo': forms.NumberInput(attrs={'class': 'form-control'}),
            'descripcion': forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
            'imagen': forms.FileInput(attrs={'class': 'form-control'}),
        }

class CategoriaForm(forms.ModelForm):
    class Meta:
        model = Categoria
        fields = ['nombre']
        labels = {
            'nombre': 'Category name'
        }
        widgets = {'nombre': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Ej. Lab'})}

class SubcategoriaForm(forms.ModelForm):
    class Meta:
        model = Subcategoria
        fields = ['categoria', 'nombre']
        labels = {
            'categoria':'Category',
            'nombre': 'Subcategory name'
        }
        widgets = {
            'categoria': forms.Select(attrs={'class': 'form-select'}),
            'nombre': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Ej. Laptops'})
        }
        
#Editar Item #1

"""
@login_required
def detalle_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    # Obtenemos solo las transacciones de ESTE ítem
    movimientos = item.transacciones.all().order_by('-fecha')
    
    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            return redirect('detalle_item', item_id=item.id)
    else:
        form = ItemForm(instance=item)
    
    return render(request, 'core/detalle_item.html', {
        'item': item,
        'form': form,
        'movimientos': movimientos
    })
"""
# Función de ayuda para verificar si es admin
def is_editor(user):
    if user.groups.filter(name='editor').exists() or user.is_superuser:
        return True
    # Si el usuario está logueado pero no es admin, lanzamos el error 403
    raise PermissionDenied

#Editar Item #2
@login_required
@user_passes_test(is_editor)
def editar_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    movimientos = item.transacciones.all().order_by('-fecha')[:10]
    
    if request.method == 'POST':
        # Pasamos la instancia actual para que Django sepa qué editar
        form = ItemForm(request.POST, request.FILES, instance=item)
        # Agregamos request.FILES para que la imagen se procese
        if form.is_valid():
            form.save()
            return redirect('detalle_item', item_id=item.id)
        else:
            # ESTO ES CLAVE: Si no se guarda, imprime los errores en tu terminal 
            # para que sepamos qué campo está protestando.
            print(form.errors) 
    else:
        form = ItemForm(instance=item)
    
    return render(request, 'core/editar_item.html', {
        'item': item,
        'form': form,
        'movimientos': movimientos
    })

@login_required
@user_passes_test(is_editor)
def crear_item(request, subcategoria_id=None):
    subcategoria_obj = None
    if subcategoria_id:
        subcategoria_obj = get_object_or_404(Subcategoria, id=subcategoria_id)

    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES)
        if form.is_valid():
            # commit=False crea el objeto pero no lo guarda todavía en la BD
            nuevo_item = form.save(commit=False)
            
            # Si venimos de una subcategoría específica, la aseguramos aquí
            if subcategoria_obj:
                nuevo_item.subcategoria = subcategoria_obj
            
            # Ahora sí, guardamos definitivamente
            nuevo_item.save()
            
            return redirect('items_por_subcategoria', 
                            categoria_id=nuevo_item.subcategoria.categoria.id, 
                            subcategoria_id=nuevo_item.subcategoria.id)
    else:
        # Pre-llenamos el formulario para que el usuario vea la subcategoría
        form = ItemForm(initial={'subcategoria': subcategoria_obj})
        
    pass
    
    return render(request, 'core/crear_item.html', {
        'form': form, 
        'subcategoria_previa': subcategoria_obj
    })
    

def cargar_subcategorias(request):
    categoria_id = request.GET.get('categoria_id')
    subcategorias = Subcategoria.objects.filter(categoria_id=categoria_id).values('id', 'nombre')
    return JsonResponse(list(subcategorias), safe=False)

def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # ESTO ES CLAVE: Forzamos el widget y las clases de nuevo
        self.fields['categoria'].widget.attrs.update({'class': 'form-select'})
        self.fields['subcategoria'].widget.attrs.update({'class': 'form-select'})
        
        if self.instance and self.instance.pk:
            # Sincronizamos la categoría con la subcategoría que ya tiene el ítem
            self.fields['categoria'].initial = self.instance.subcategoria.categoria.id

@login_required
@user_passes_test(is_editor)
def gestionar_categorias(request):
    categorias = Categoria.objects.all().prefetch_related('subcategorias')
    form_cat = CategoriaForm()
    form_sub = SubcategoriaForm()

    if request.method == 'POST':
        if 'btn_cat' in request.POST:
            form_cat = CategoriaForm(request.POST)
            if form_cat.is_valid():
                form_cat.save()
                return redirect('gestionar_categorias')
        elif 'btn_sub' in request.POST:
            form_sub = SubcategoriaForm(request.POST)
            if form_sub.is_valid():
                form_sub.save()
                return redirect('gestionar_categorias')

    return render(request, 'core/gestionar_categorias.html', {
        'categorias': categorias,
        'form_cat': form_cat,
        'form_sub': form_sub
    })

@login_required
def editar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            return redirect('gestionar_categorias')
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, 'core/editar_clasificacion.html', {'form': form, 'titulo': 'Editar Categoría'})

@login_required
def editar_subcategoria(request, pk):
    subcategoria = get_object_or_404(Subcategoria, pk=pk)
    if request.method == 'POST':
        form = SubcategoriaForm(request.POST, instance=subcategoria)
        if form.is_valid():
            form.save()
            return redirect('gestionar_categorias')
    else:
        form = SubcategoriaForm(instance=subcategoria)
    return render(request, 'core/editar_clasificacion.html', {'form': form, 'titulo': 'Editar Subcategoría'})

@login_required
def home_categorias(request):
    # Traemos las categorías. El conteo de subcategorías lo hace el template
    # con cat.subcategorias.count, que es mucho más seguro.
    categorias = Categoria.objects.all()
    return render(request, 'core/home_categorias.html', {'categorias': categorias})

@login_required
def items_por_categoria(request, categoria_id, subcategoria_id=None):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    
    # Ordenamos las subcategorías alfabéticamente por nombre
    subcategorias = Subcategoria.objects.filter(categoria=categoria)
    
    if subcategoria_id:
        subcategoria_activa = get_object_or_404(Subcategoria, id=subcategoria_id)
    else:
        # Al estar ordenadas arriba, .first() ahora devolverá la primera alfabéticamente (ej: "A...")
        subcategoria_activa = subcategorias.first()

    items = Item.objects.filter(subcategoria=subcategoria_activa)
    
    return render(request, 'core/items_por_categoria.html', {
        'categoria': categoria,
        'subcategorias': subcategorias,
        'subcategoria_activa': subcategoria_activa,
        'items': items,
    })

@login_required
def registrar_movimiento(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        cantidad = int(request.POST.get('cantidad'))
        
        # --- VALIDACIÓN DE SEGURIDAD ---
        if tipo == 'SALIDA' and cantidad > item.stock:
            messages.error(request, f"⚠️ Error: No puedes retirar {cantidad} unidades. El stock actual es de solo {item.stock}.")
            return redirect('detalle_item', item_id=item.id)
        # -------------------------------

        # Si pasa la validación, procedemos a guardar
        stock_previo = item.stock
        if tipo == 'ENTRADA':
            item.stock += cantidad
        elif tipo == 'SALIDA':
            item.stock -= cantidad
            
        item.save()
        
        # Registrar la transacción en el historial
        Transaccion.objects.create(
            item=item,
            tipo=tipo,
            cantidad=cantidad,
            stock_previo=stock_previo,
            stock_nuevo=item.stock,
            usuario=request.user.username
        )
        
        messages.success(request, f"Stock actualizado: {item.nombre} ahora tiene {item.stock} unidades.")
        return redirect(request.META.get('HTTP_REFERER', 'home_categorias'))


@login_required
def despacho_multiple(request):
    items = Item.objects.all().order_by('nombre')
    
    if request.method == 'POST':
        item_ids = request.POST.getlist('item_id[]')
        cantidades = request.POST.getlist('cantidad[]')
        errores = []

        for i_id, cant in zip(item_ids, cantidades):
            if i_id and cant:
                item = get_object_or_404(Item, id=i_id)
                cantidad = int(cant)
                
                if cantidad > item.stock:
                    errores.append(f"{item.nombre} (Stock insuficiente: {item.stock})")
                    continue # Salta este item y sigue con el próximo

                # Lógica de descuento
                stock_previo = item.stock
                item.stock -= cantidad
                item.save()
                
                Transaccion.objects.create(
                    item=item, tipo='SALIDA', cantidad=cantidad,
                    stock_previo=stock_previo, stock_nuevo=item.stock,
                    usuario=request.user.username
                )

        if errores:
            messages.warning(request, f"Se procesaron los cambios, pero hubo problemas con: {', '.join(errores)}")
        else:
            messages.success(request, "Todo el despacho se registró correctamente.")
            
        return redirect('home_categorias')

    return render(request, 'core/despacho_multiple.html', {'items': items})

@login_required
def reporte_subcategoria(request, sub_id):
    sub = get_object_or_404(Subcategoria, id=sub_id)
    items = Item.objects.filter(subcategoria=sub).order_by('nombre')
    fecha = datetime.now()
    return render(request, 'core/reporte_pdf.html', {'items': items, 'subcategoria': sub, 'fecha': fecha})

@login_required
def panel_reportes(request):
    # 'subcategorias' debe coincidir con el related_name del modelo
    categorias = Categoria.objects.prefetch_related('subcategorias').all()
    return render(request, 'core/panel_reportes.html', {'categorias': categorias})

@login_required
def detalle_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    return render(request, 'core/detalle_item.html', {'item': item})

@login_required
def dashboard(request):
    return render(request, 'core/dashboard.html')

@login_required
def pagina_exportar_filtros(request):
    categorias = Categoria.objects.all()
    subcategorias = Subcategoria.objects.all()
    return render(request, 'core/exportar_filtros.html', {
        'categorias': categorias,
        'subcategorias': subcategorias
    })

@login_required
def exportar_inventario_excel(request):
    # 1. Obtener parámetros de los filtros
    categoria_id = request.GET.get('categoria')
    subcategoria_id = request.GET.get('subcategoria')

    # 2. Empezar con todos los items (QuerySet base)
    items = Item.objects.all().select_related('subcategoria__categoria')

    # 3. Aplicar filtros dinámicamente
    if categoria_id and categoria_id != 'todas':
        items = items.filter(subcategoria__categoria_id=categoria_id)
    
    if subcategoria_id and subcategoria_id != 'todas':
        items = items.filter(subcategoria_id=subcategoria_id)

    # --- AQUÍ ESTABA EL ERROR: No vuelvas a llamar a Item.objects.all() ---

    # 4. Crear el libro y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Clínica"

    # 5. Estilo de encabezado (opcional, para que combine con tu azul #012768)
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="012768", end_color="012768", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Definir el encabezado
    columns = ['Nombre', 'Categoría', 'Subcategoría', 'Stock Actual', 'Stock Mínimo', 'Estado']
    ws.append(columns)

    # Aplicar estilo al encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 6. Escribir los datos (usando el QuerySet ya filtrado)
    for item in items:
        # Lógica para el estado
        estado = "OK"
        if item.stock <= 0:
            estado = "AGOTADO"
        elif item.stock <= item.stock_minimo:
            estado = "CRÍTICO"

        ws.append([
            item.nombre,
            str(item.subcategoria.categoria.nombre if item.subcategoria else "S/C"),
            str(item.subcategoria.nombre if item.subcategoria else "S/S"),
            item.stock,
            item.stock_minimo,
            estado
        ])

    # 7. Ajustar el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    # 8. Preparar la respuesta del navegador
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="inventario_clinica_filtrado.xlsx"'
    
    wb.save(response)
    return response

@login_required
def exportar_subcategoria_excel(request, subcategoria_id):
    # 1. Obtener los datos
    subcategoria = Subcategoria.objects.get(pk=subcategoria_id)
    items = Item.objects.filter(subcategoria=subcategoria)

    # 2. Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # 3. Estilo para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="012768", end_color="012768", fill_type="solid") # Tu azul corporativo
    
    # Definir títulos de columnas
    columns = ['Nombre del Insumo', 'Stock Actual', 'Unidad', 'Última Actualización']
    ws.append(columns)

    # Aplicar estilos a los encabezados
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 4. Cargar los datos de los ítems
    for item in items:
        ws.append([
            item.nombre,
            item.stock,
        ])

    # 5. Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except: pass
        ws.column_dimensions[column].width = max_length + 5

    # 6. Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Inventario_{subcategoria.nombre}.xlsx"'
    
    wb.save(response)
    return response